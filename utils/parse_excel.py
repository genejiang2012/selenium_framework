# encoding= utf-8

import openpyxl
from openpyxl.styles import Border, Side, Font
import time


class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)

        # 设置 字体 的 颜色 # 颜色 对应 的 RGB 值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        # 将Excel文件加载到内存， 并获取其 workbook 对象
        self.workbook = openpyxl.load_workbook(excelPathAndName)
        self.excelFile = excelPathAndName

        return self.workbook

    def getSheetByName(self, sheetName):
        # 根据 sheet 名 获取 该 sheet 对象
        try:
            print(sheetName)
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        # 根据 sheet 的 索引 号 获取 该 sheet 对象
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e

        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self, sheet):
        # 获取 sheet 中有 数据区 域 的 结束 行号
        return sheet.max_row

    def getColsNumber(self, sheet):
        # 获取 sheet 中有 数据区 域 的 结束 列 号
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取 sheet 中有 数据区 域 的 开始 行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        # 获取 sheet 中有 数据区 域 的 开始 列 号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中某一行，返回的是这一行 所有 的数据 内容 组成 的 tuple
        # 下标 从 1 开始， sheet.rows［1］ 表示 第一 行
        try:
            return list(sheet.rows)[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取 sheet 中 某一 列，返回 的是这一列所有的数据内容组成 tuple，
        # 下标 从 1 开始， sheet. columns［ 1］ 表示 第 一列
        try:
            return list(sheet.columns)[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate=None,
                       rowNo=None, colsNo=None):
        # 根据 单元格 所在 的 位置 索引 获取 该 单元格 中的 值，
        # 下标 从 1 开始，sheet.cell(row=1, column=1).value
        # 表示excel中第一行第一列的值
        if coordinate is not None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellOfObject(self, sheet, coordinate=None,
                        rowNo=None, colsNo=None):
        # 获取 某个 单元格 的 对象， 可以 根据 单元格 所在 位置 的 数字 索引，
        # 也可以 直接 根据 Excel 中 单元格 的 编码 及 坐标
        # 如 getCellObject( sheet, coordinate = 'A1') or
        # getCellObject( sheet, rowNo = 1, colsNo = 2)
        if coordinate is not None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception(" Insufficient Coordinates of cell !")

    def writeCell(self, sheet, content, coordinate=None, rowNo=None,
                  colsNo=None, style=None):
        # 根据 单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        # 下标从 1开始，参数 style 表示 字体 的 颜色 的 名字，比如 red， green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content

                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(
                        color=self.RGBDict[style])

                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(
                        color=self.RGBDict[style])

                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception(" Insufficient Coordinates of cell !")

    def writeCellCurrentTime(self, sheet, coordinate = None,
                             rowNo = None, colsNo = None):
        # 写入当前的时间，下标从1开始
        now = int( time. time()) #显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime

                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = currentTime

                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception(" Insufficient Coordinates of cell !")



if __name__ == "__main__":
    pe = ParseExcel()
    pe.loadWorkBook(u'D:\\Git_genejiang2012\\KeyWord_Data_Driven_Framework\\testdata\\example.xlsx')
    print(pe.getSheetByName("Login").title)
    print(type(pe.getSheetByName("Sheet1")))
    print(pe.getSheetByIndex(1).title)
    sheet_obj = pe.getSheetByIndex(0)
    print(type(sheet_obj))
    print(pe.getStartColNumber(sheet_obj))
    print(pe.getStartColNumber(sheet_obj))
    print(pe.getRowsNumber(sheet_obj))
    print(pe.getColsNumber(sheet_obj))
    print(pe.getRow(sheet_obj, 2))
    print(pe.getColumn(sheet_obj, 2))
    rows = pe.getRow(sheet_obj, 1)
    for i in rows:
        print(i.value)
    print(pe.getCellOfObject(sheet_obj, rowNo=1, colsNo=1))
    print(pe.getCellOfValue(sheet_obj, rowNo=1, colsNo=1))
    pe.writeCell(sheet_obj, u'I love my country', rowNo=1, colsNo=4,
                 style='green')
    pe.writeCellCurrentTime(sheet_obj, rowNo=5, colsNo=4)
